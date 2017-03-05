from openpyxl import Workbook
import totalwords

lexicalEntries = 'lexicalEntries'

entries = 'entries'

senses = 'senses'

definitions = 'definitions'

examples = 'examples'

etymologies = 'etymologies'

class excelCreator:

    wb = Workbook()

    ws = wb.active

    def create(self, words, fileName):

        results = []

        for word in words:

            item = totalwords.instance().getWord(word)

            if item != 0:

                results.append(item)

        self.writeToExcel(self, results, fileName)

    @classmethod
    def writeToExcel(cls, words, fileName):

        nIndex = 0

        for word in words:

            nIndex += + 1

            cls.ws['A' + str(nIndex)] = word['id']

            cls.ws['B' + str(nIndex)], cls.ws['C' + str(nIndex)], cls.ws['D' + str(nIndex)] = word['def'], word['exam'], word['ety']

        cls.wb.save(fileName)

__intance = 0

def instance():

    global  __intance

    if __intance == 0:

        __intance = excelCreator()

    return __intance





